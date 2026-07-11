import os
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def generate_csv_report(predictions_list):
    """
    Generates a CSV string buffer of prediction history.
    """
    output = BytesIO()
    # Write to BytesIO as string, then get bytes
    wrapper = BytesIO()
    # Python csv writer requires text mode, so we use string buffer and convert at the end
    text_buffer = []
    
    header = ['ID', 'Date & Time', 'Image Name', 'Predicted Class', 'Confidence (%)', 'Latency (s)', 'Recycling Instructions', 'Environmental Impact']
    text_buffer.append(header)
    
    for row in predictions_list:
        text_buffer.append([
            row.get('id', ''),
            row.get('created_at', ''),
            row.get('image_name', ''),
            row.get('predicted_class', ''),
            f"{row.get('confidence', 0):.2f}",
            f"{row.get('processing_time', 0):.4f}",
            row.get('recycling_instructions', ''),
            row.get('environmental_tip', '')
        ])
        
    # Write to bytes
    io_string = BytesIO()
    # we can construct the CSV string directly
    csv_str = ""
    for r in text_buffer:
        # Simple escape quotes and wrap in quotes if contains comma
        escaped_r = []
        for col in r:
            val = str(col).replace('"', '""')
            if ',' in val or '\n' in val or '"' in val:
                escaped_r.append(f'"{val}"')
            else:
                escaped_r.append(val)
        csv_str += ",".join(escaped_r) + "\r\n"
        
    return csv_str.encode('utf-8')

def generate_xlsx_report(predictions_list):
    """
    Generates an Excel binary buffer of prediction history.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Waste Classification Log"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='2E7D32')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Segoe UI', size=10)
    
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    zebra_fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
    
    border_thin = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Set title block
    ws['A1'] = "Smart Waste Management - Classification History Report"
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30
    
    # Set headers
    headers = ['ID', 'Date & Time', 'Image Name', 'Predicted Class', 'Confidence (%)', 'Latency (s)', 'Recycling Instructions', 'Environmental Impact']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
    ws.row_dimensions[3].height = 25
    
    # Populate data
    for row_idx, data in enumerate(predictions_list, 4):
        vals = [
            data.get('id', ''),
            data.get('created_at', ''),
            data.get('image_name', ''),
            data.get('predicted_class', ''),
            round(data.get('confidence', 0), 2),
            round(data.get('processing_time', 0), 4),
            data.get('recycling_instructions', ''),
            data.get('environmental_tip', '')
        ]
        
        is_even = (row_idx % 2 == 0)
        
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_thin
            
            # Formatting and Alignment
            if col_idx in [1, 2, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx in [5, 6]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
            if is_even:
                cell.fill = zebra_fill
                
        ws.row_dimensions[row_idx].height = 20
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        # Don't size column widths based on title A1
        for cell in col:
            if cell.row != 1 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Specific widths for descriptions
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 35
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_pdf_report(prediction_record):
    """
    Generates a high-quality PDF report for a single waste classification record.
    """
    buffer = BytesIO()
    
    # Page setup
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette
    primary_color = colors.HexColor('#1B5E20') # Deep Green
    secondary_color = colors.HexColor('#4CAF50') # Light Green
    text_dark = colors.HexColor('#212121')
    bg_light = colors.HexColor('#F1F8E9')
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=primary_color,
        spaceAfter=15,
        alignment=1 # Center
    )
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=primary_color,
        spaceBefore=10,
        spaceAfter=6,
        borderPadding=2
    )
    
    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=text_dark,
        leading=14
    )
    
    body_bold_style = ParagraphStyle(
        'DocBodyBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    alert_style = ParagraphStyle(
        'DocAlert',
        parent=body_style,
        textColor=colors.HexColor('#33691E')
    )

    story = []
    
    # Header Logo / Title
    story.append(Paragraph("Smart Waste Classification Report", title_style))
    story.append(Paragraph("AI-powered classification and ecological disposal recommendation for sustainable waste management.", ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=1, textColor=colors.HexColor('#757575'), spaceAfter=20)))
    
    # Metadata Table
    meta_data = [
        [Paragraph("Report ID:", body_bold_style), Paragraph(f"WMS-{prediction_record.get('id', 'N/A')}", body_style),
         Paragraph("Classification Date:", body_bold_style), Paragraph(str(prediction_record.get('created_at', 'N/A')), body_style)],
        [Paragraph("Image File Name:", body_bold_style), Paragraph(prediction_record.get('image_name', 'N/A'), body_style),
         Paragraph("Inference Latency:", body_bold_style), Paragraph(f"{prediction_record.get('processing_time', 0.0):.4f} seconds", body_style)]
    ]
    
    meta_table = Table(meta_data, colWidths=[1.5*inch, 2.0*inch, 1.7*inch, 2.0*inch])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FAFAFA'))
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 20))
    
    # Results Section
    story.append(Paragraph("AI Model Classification Results", section_title_style))
    
    res_data = [
        [Paragraph("Predicted Material Class", body_bold_style), Paragraph("Model Confidence", body_bold_style)],
        [Paragraph(f"<font size=16 color='#1B5E20'><b>{prediction_record.get('predicted_class', 'Unknown')}</b></font>", body_style),
         Paragraph(f"<font size=16 color='#1B5E20'><b>{prediction_record.get('confidence', 0.0):.2f}%</b></font>", body_style)]
    ]
    res_table = Table(res_data, colWidths=[3.6*inch, 3.6*inch])
    res_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, primary_color),
        ('PADDING', (0,0), (-1,-1), 12),
        ('BACKGROUND', (0,0), (-1,-1), bg_light)
    ]))
    story.append(res_table)
    story.append(Spacer(1, 20))
    
    # Images Section (If images exist, add them)
    # Note: image_path is relative, so we need to construct the absolute path to load
    # Let's locate the project root and load the image if it's there
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    orig_path = prediction_record.get('original_image_path', '')
    heat_path = prediction_record.get('heatmap_image_path', '')
    
    # Strip leading slash if any
    if orig_path.startswith('/'): orig_path = orig_path[1:]
    if heat_path and heat_path.startswith('/'): heat_path = heat_path[1:]
    
    abs_orig_path = os.path.join(project_root, orig_path)
    abs_heat_path = os.path.join(project_root, heat_path) if heat_path else None
    
    images_table_data = []
    col_widths = []
    
    if os.path.exists(abs_orig_path):
        try:
            img_w, img_h = 2.5*inch, 2.5*inch
            # Convert original image flowable
            orig_flowable = Image(abs_orig_path, width=img_w, height=img_h)
            
            if abs_heat_path and os.path.exists(abs_heat_path):
                # We have Grad-CAM! Put them side by side
                heat_flowable = Image(abs_heat_path, width=img_w, height=img_h)
                images_table_data.append([orig_flowable, heat_flowable])
                images_table_data.append([Paragraph("<b>Original Uploaded Image</b>", body_style), Paragraph("<b>Explainable AI (Grad-CAM Activation Grid)</b>", body_style)])
                col_widths = [3.6*inch, 3.6*inch]
            else:
                images_table_data.append([orig_flowable])
                images_table_data.append([Paragraph("<b>Original Uploaded Image</b>", body_style)])
                col_widths = [7.2*inch]
        except Exception as e:
            print(f"Error loading image in PDF report: {e}")
            
    if images_table_data:
        img_table = Table(images_table_data, colWidths=col_widths)
        img_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(Paragraph("Visual Evidence & Heatmap Overlay", section_title_style))
        story.append(img_table)
        story.append(Spacer(1, 20))
        
    # Recommendations & Impact Details
    story.append(Paragraph("Recycling & Ecological Recommendations", section_title_style))
    
    rec_details = [
        [Paragraph("<b>Recycling Instructions:</b>", body_bold_style), Paragraph(prediction_record.get('recycling_instructions', 'N/A'), body_style)],
        [Paragraph("<b>Disposal Advice:</b>", body_bold_style), Paragraph(prediction_record.get('recycling_instructions', 'N/A'), body_style)], # Fallback if advice blank
        [Paragraph("<b>Environmental Impact Fact:</b>", body_bold_style), Paragraph(prediction_record.get('environmental_tip', 'N/A'), alert_style)]
    ]
    
    rec_table = Table(rec_details, colWidths=[1.8*inch, 5.4*inch])
    rec_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor('#EEEEEE')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#F9FBE7'))
    ]))
    story.append(rec_table)
    story.append(Spacer(1, 20))
    
    # XAI explanation
    story.append(Paragraph("Explainable AI (XAI) Model Explanation", section_title_style))
    story.append(Paragraph(
        prediction_record.get('explain_text', 
                              "The deep neural network MobileNetV2 was fine-tuned using transfer learning to map global features from layers of Convolution, Batch Normalization, and ReLU activations. The Grad-CAM heatmap identifies region-specific patterns of light-diffraction and edge density corresponding to features learned from the dataset (e.g. sharp edges, curved reflections, or material textures)."),
        body_style
    ))
    
    story.append(Spacer(1, 30))
    story.append(Paragraph("<i>Thank you for helping keep our planet clean!</i>", ParagraphStyle('Footer', parent=styles['Normal'], alignment=1, textColor=primary_color, fontName='Helvetica-Oblique', fontSize=10)))
    
    # Build Document
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
